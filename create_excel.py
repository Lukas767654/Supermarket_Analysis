#!/usr/bin/env python3
"""
Excel Creator - Erstellt eine moderne Excel-Datei aus allen CSV-Dateien
"""

import pandas as pd
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference
import glob

def create_modern_excel():
    """Erstelle eine moderne Excel-Datei aus allen CSV-Dateien"""
    
    print("📊 Creating Modern Excel File")
    print("=" * 40)
    
    # Finde alle CSV-Dateien
    csv_folder = "Analysis/csv_exports"
    csv_files = glob.glob(f"{csv_folder}/*.csv")
    
    print(f"Found {len(csv_files)} CSV files:")
    for file in csv_files:
        print(f"  📄 {os.path.basename(file)}")
    
    if not csv_files:
        print("❌ No CSV files found!")
        return
    
    # Erstelle Excel Workbook
    wb = Workbook()
    
    # Entferne das Standard-Sheet
    wb.remove(wb.active)
    
    # Definiere moderne Styling
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    data_font = Font(name='Segoe UI', size=10)
    border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    for csv_file in sorted(csv_files):
        filename = os.path.basename(csv_file)
        sheet_name = filename.replace('.csv', '').replace('_', ' ').title()
        
        print(f"\n📝 Processing {filename} -> {sheet_name}")
        
        # Lade CSV
        try:
            df = pd.read_csv(csv_file)
            print(f"   📊 {len(df)} rows, {len(df.columns)} columns")
            
            # Erstelle neues Sheet
            ws = wb.create_sheet(title=sheet_name)
            
            # Füge Daten hinzu
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Style Header
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            
            # Style Data
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = border
                    
                    # Alternating row colors
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set width with reasonable limits
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Add conditional formatting for numeric columns
            for col_idx, column_name in enumerate(df.columns, 1):
                if df[column_name].dtype in ['int64', 'float64']:
                    column_letter = ws.cell(row=1, column=col_idx).column_letter
                    rule = ColorScaleRule(
                        start_type='min', start_color='E8F5E8',
                        mid_type='percentile', mid_value=50, mid_color='FFF2CC',
                        end_type='max', end_color='F8CECC'
                    )
                    ws.conditional_formatting.add(f'{column_letter}2:{column_letter}{len(df)+1}', rule)
            
            print(f"   ✅ Sheet '{sheet_name}' created successfully")
            
        except Exception as e:
            print(f"   ❌ Error processing {filename}: {e}")
    
    # Füge Summary Sheet hinzu
    create_summary_sheet(wb, csv_files)
    
    # Speichere Excel-Datei
    output_file = "Analysis/Supermarket_Analysis_Complete.xlsx"
    wb.save(output_file)
    
    print(f"\n🎉 SUCCESS!")
    print(f"📊 Modern Excel file created: {output_file}")
    print(f"📁 Contains {len(wb.worksheets)} sheets with professional formatting")
    
    return output_file

def create_summary_sheet(wb, csv_files):
    """Erstelle ein Summary Sheet mit Übersicht"""
    
    print(f"\n📋 Creating Summary Sheet")
    
    # Erstelle Summary Sheet
    ws = wb.create_sheet(title="📊 Overview", index=0)
    
    # Styling
    title_font = Font(name='Segoe UI', size=16, bold=True, color='2E86AB')
    header_font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    data_font = Font(name='Segoe UI', size=10)
    border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    # Titel
    ws['A1'] = '🏪 Supermarket Analysis - Complete Dataset'
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    
    # Dataset Overview
    ws['A3'] = 'Dataset Overview'
    ws['A3'].font = Font(name='Segoe UI', size=14, bold=True)
    
    # Headers
    headers = ['Sheet Name', 'Rows', 'Columns', 'Description', 'Key Metrics']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Data für jedes Sheet
    sheet_descriptions = {
        'eye_level_analysis': 'Eye-level positioning and shelf analysis',
        'brand_classification': 'Brand origin classification with confidence scores', 
        'cjmore_private_brands': 'Private brand analysis for CJ More',
        'thai_vs_international': 'Comparison between Thai and international brands'
    }
    
    row = 6
    total_products = 0
    
    for csv_file in sorted(csv_files):
        filename = os.path.basename(csv_file).replace('.csv', '')
        
        try:
            df = pd.read_csv(csv_file)
            sheet_name = filename.replace('_', ' ').title()
            
            # Sheet info
            ws.cell(row=row, column=1, value=sheet_name).font = data_font
            ws.cell(row=row, column=2, value=len(df)).font = data_font  
            ws.cell(row=row, column=3, value=len(df.columns)).font = data_font
            ws.cell(row=row, column=4, value=sheet_descriptions.get(filename, 'Data analysis')).font = data_font
            
            # Key metrics
            if 'category' in df.columns:
                categories = df['category'].nunique()
                ws.cell(row=row, column=5, value=f'{categories} categories').font = data_font
            else:
                ws.cell(row=row, column=5, value='N/A').font = data_font
            
            # Styling
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
                if row % 2 == 0:
                    ws.cell(row=row, column=col).fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            
            total_products += len(df)
            row += 1
            
        except Exception as e:
            print(f"   ⚠️ Could not process {filename}: {e}")
    
    # Summary stats
    ws[f'A{row+2}'] = 'Summary Statistics'
    ws[f'A{row+2}'].font = Font(name='Segoe UI', size=14, bold=True)
    
    ws[f'A{row+4}'] = f'📦 Total Products Analyzed: {total_products:,}'
    ws[f'A{row+4}'].font = Font(name='Segoe UI', size=11, bold=True)
    
    ws[f'A{row+5}'] = f'📊 Analysis Sheets: {len(csv_files)}'
    ws[f'A{row+5}'].font = Font(name='Segoe UI', size=11)
    
    ws[f'A{row+6}'] = f'📅 Generated: {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")}'
    ws[f'A{row+6}'].font = Font(name='Segoe UI', size=11)
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 20
    
    print(f"   ✅ Summary sheet created with overview of {len(csv_files)} datasets")

def main():
    """Hauptfunktion"""
    
    print("🏪 SUPERMARKET ANALYSIS - EXCEL CREATOR")
    print("=" * 50)
    
    # Prüfe ob CSV-Ordner existiert
    if not os.path.exists("Analysis/csv_exports"):
        print("❌ CSV exports folder not found!")
        return
    
    # Erstelle Excel-Datei
    output_file = create_modern_excel()
    
    if output_file:
        print(f"\n📎 Ready to use: {output_file}")
        print("🎨 Features:")
        print("  ✅ Modern, minimalist design")
        print("  ✅ Professional color scheme") 
        print("  ✅ Auto-sized columns")
        print("  ✅ Frozen headers")
        print("  ✅ Alternating row colors")
        print("  ✅ Conditional formatting")
        print("  ✅ Summary overview sheet")

if __name__ == "__main__":
    main()