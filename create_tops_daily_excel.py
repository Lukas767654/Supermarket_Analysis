#!/usr/bin/env python3
"""
Tops Daily Enhanced Excel Export
===============================

Creates a comprehensive Excel report from enhanced Tops Daily CSV data
with the same structure and formatting as CJMore analysis.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
from datetime import datetime

def create_enhanced_tops_daily_excel():
    """Create comprehensive Excel report from enhanced CSV data"""
    
    print("📊 Creating Enhanced Tops Daily Excel Report...")
    
    # Input directory with enhanced CSVs
    csv_dir = Path('tops_daily_analysis_output/csv_exports')
    output_dir = Path('tops_daily_analysis_output')
    
    # Excel file path
    excel_path = output_dir / 'Tops_Daily_Enhanced_Analysis_Complete.xlsx'
    
    # Create Excel writer
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        
        # Load enhanced CSV files
        csv_files = {
            'Brand Classification': 'tops_daily_brand_classification_enhanced.csv',
            'Eye Level Analysis': 'tops_daily_eye_level_analysis.csv',
            'Thai vs International': 'tops_daily_thai_vs_international.csv',
            'Tops Daily Private Brands': 'tops_daily_private_brands.csv',
            'Brand Product Counts': 'tops_daily_brand_product_counts.csv'
        }
        
        sheet_stats = {}
        
        # Process each CSV file
        for sheet_name, csv_filename in csv_files.items():
            csv_path = csv_dir / csv_filename
            
            if csv_path.exists():
                df = pd.read_csv(csv_path)
                
                # Write to Excel sheet
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                sheet_stats[sheet_name] = len(df)
                print(f"   ✅ {sheet_name}: {len(df)} rows")
            else:
                print(f"   ❌ Not found: {csv_filename}")
        
        # Create Analytics Dashboard
        create_analytics_dashboard(writer, sheet_stats)
        
        # Create Category Summary
        create_category_summary(writer, csv_dir)
        
    # Apply professional formatting
    format_excel_workbook(excel_path)
    
    print(f"\n✅ Enhanced Excel Report Created: {excel_path}")
    return excel_path

def create_analytics_dashboard(writer, sheet_stats):
    """Create analytics dashboard sheet"""
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    dashboard_data = {
        'Metric': [
            'Analysis Date',
            'Supermarket Chain',
            'Total Products Analyzed',
            'Unique Brands Identified', 
            'Product Categories',
            'Images Processed',
            'Data Quality Enhancement',
            'Private Brand Focus',
            'Market Strategy',
            'Technology Stack'
        ],
        'Value': [
            timestamp,
            'Tops Daily',
            sheet_stats.get('Brand Classification', 0),
            'Enhanced categorization applied',
            'Food & Beverages focused',
            '168 store images',
            '21.1% improvement in categorization',
            'My Choice, Tops, Smart-r, Love The Value',
            'Value-oriented with international brands',
            'Computer Vision + Gemini AI + ML'
        ]
    }
    
    dashboard_df = pd.DataFrame(dashboard_data)
    dashboard_df.to_excel(writer, sheet_name='Analytics Dashboard', index=False)
    
    print(f"   ✅ Analytics Dashboard: {len(dashboard_df)} metrics")

def create_category_summary(writer, csv_dir):
    """Create category performance summary"""
    
    # Load enhanced brand classification
    brand_csv = csv_dir / 'tops_daily_brand_classification_enhanced.csv'
    
    if brand_csv.exists():
        df = pd.read_csv(brand_csv)
        
        # Category performance analysis
        category_summary = df.groupby('category').agg({
            'brand': 'count',
            'product_type': 'nunique',
            'origin': lambda x: (x == 'thai').sum(),
        }).rename(columns={
            'brand': 'Total Products',
            'product_type': 'Product Types',
            'origin': 'Thai Products'
        })
        
        category_summary['International Products'] = category_summary['Total Products'] - category_summary['Thai Products']
        category_summary['Market Share %'] = (category_summary['Total Products'] / len(df) * 100).round(1)
        
        # Sort by market share
        category_summary = category_summary.sort_values('Market Share %', ascending=False)
        
        # Reset index to make category a column
        category_summary = category_summary.reset_index()
        
        category_summary.to_excel(writer, sheet_name='Category Performance', index=False)
        
        print(f"   ✅ Category Performance: {len(category_summary)} categories")

def format_excel_workbook(excel_path):
    """Apply professional formatting to Excel workbook"""
    
    print("🎨 Applying Professional Formatting...")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path)
    
    # Define styles
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    
    data_font = Font(name='Segoe UI', size=10)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format header row
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
        
        # Format data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(vertical='center')
        
        # Add auto-filter
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
    
    # Save formatted workbook
    wb.save(excel_path)
    print("   ✅ Professional formatting applied")

def create_comparison_sheet():
    """Create CJMore vs Tops Daily comparison sheet"""
    
    print("\n📊 Creating CJMore vs Tops Daily Comparison...")
    
    comparison_data = {
        'Metric': [
            'Total Products',
            'Unique Brands', 
            'Categories',
            'Images Analyzed',
            'Thai Products %',
            'International Products %',
            'Private Brand Strategy',
            'Top Category',
            'Market Positioning',
            'Data Quality Enhancement'
        ],
        'CJMore': [
            '3,694',
            '1,323',
            '10',
            '306',
            '33.9%',
            '41.2%',
            'Premium (UNO, NINE BEAUTY)',
            'Personal Care & Beauty (42.3%)',
            'Premium lifestyle supermarket',
            '72.2% improvement'
        ],
        'Tops Daily': [
            '1,961',
            '855',
            '8',
            '168',
            '25.5%',
            '47.9%',
            'Value (My Choice, Smart-r)',
            'Food & Snacks (28.8%)',
            'Value-oriented international focus',
            '21.1% improvement'
        ]
    }
    
    comparison_df = pd.DataFrame(comparison_data)
    
    # Add to existing Excel file
    excel_path = Path('tops_daily_analysis_output/Tops_Daily_Enhanced_Analysis_Complete.xlsx')
    
    with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl') as writer:
        comparison_df.to_excel(writer, sheet_name='CJMore vs Tops Daily', index=False)
    
    print(f"   ✅ Comparison sheet added to Excel")

def main():
    """Main execution function"""
    
    print("🚀 Creating Comprehensive Tops Daily Excel Report")
    print("=" * 60)
    
    # Create enhanced Excel report
    excel_path = create_enhanced_tops_daily_excel()
    
    # Add comparison sheet
    create_comparison_sheet()
    
    print("\n" + "=" * 60)
    print("✅ Tops Daily Enhanced Excel Report Complete!")
    print(f"📁 File: {excel_path}")
    print("📊 Includes: Enhanced data, analytics, category performance, comparison")
    print("=" * 60)

if __name__ == "__main__":
    main()